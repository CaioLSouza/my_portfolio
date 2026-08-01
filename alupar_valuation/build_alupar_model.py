"""Monta o modelo de valuation da Alupar (ALUP11) em Excel.

Todo o workbook é dirigido por fórmulas a partir das abas `Inputs` e `Dados de origem`.
Nenhum número calculado é gravado como valor: as únicas células com número fixo são
as premissas (amarelo) e os dados públicos coletados (azul).

    python build_alupar_model.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as L
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart import LineChart, Reference

# ------------------------------------------------------------------ parâmetros do build
ANO0, ANOF = 2026, 2080                      # horizonte longo: toda safra de capex vive até o fim
ANOS = list(range(ANO0, ANOF + 1))
N = len(ANOS)

SH = dict(read='Read me', inp='Inputs', src='Dados de origem', tr='Transmissão',
          ge='Geração', cons='Consolidado', val='Valuation', sens='Sensibilidade')

A = 'Arial'
F_T   = Font(name=A, size=13, bold=True, color='1A2226')
F_H   = Font(name=A, size=9,  bold=True, color='FFFFFF')
F_IN  = Font(name=A, size=11, bold=True, color='0000FF')   # premissa editável
F_SRC = Font(name=A, size=10, color='0000FF')              # dado público coletado
F_C   = Font(name=A, size=10, color='000000')              # fórmula
F_K   = Font(name=A, size=10, bold=True, color='000000')
F_N   = Font(name=A, size=9,  color='595959')
FH    = PatternFill('solid', fgColor='2F6F74')
FIN   = PatternFill('solid', fgColor='FFFF00')
FKEY  = PatternFill('solid', fgColor='EAF2F2')
FWARN = PatternFill('solid', fgColor='FFF4E0')
BOX   = Border(*[Side(style='thin', color='BFBFBF')] * 4)
PCT, PCT1, NUM, NUM0, DT, INT, PRC = '0.00%', '0.0%', '#,##0.0', '#,##0', 'yyyy-mm-dd', '0', 'R$ #,##0.00'
WRAP = Alignment(wrap_text=True, vertical='top')

wb = Workbook()
wb.remove(wb.active)


def head(ws, row, labels, widths):
    for j, s in enumerate(labels, 1):
        c = ws.cell(row, j, s)
        c.font, c.fill = F_H, FH
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 34
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[L(j)].width = w


def put(ws, r, c, v, nf=None, font=F_C, fill=None):
    cell = ws.cell(r, c)
    if v is not None:
        cell.value = v
    if nf:
        cell.number_format = nf
    cell.font = font
    if fill:
        cell.fill = fill
    return cell


def name(n, sheet, ref):
    wb.defined_names.add(DefinedName(n, attr_text=f"'{sheet}'!{ref}"))


# ============================================================ Dados de origem (azul)
ws = wb.create_sheet(SH['src'])
ws.cell(1, 1, 'Dados de origem — números públicos coletados').font = F_T
ws.cell(2, 1, 'Azul = coletado de fonte pública, com a fonte na última coluna. Não editar sem trocar a fonte. '
              'Tudo que a aba Inputs assume a partir daqui é premissa, não dado.').font = F_N
for col, w in [('A', 3), ('B', 44), ('C', 16), ('D', 12), ('E', 64)]:
    ws.column_dimensions[col].width = w
head(ws, 4, ['', 'Item', 'Valor', 'Unidade', 'Fonte'], [3, 44, 16, 12, 64])

SRC_ROWS = [
    ('MERCADO', None, None, None),
    ('Cotação ALUP11', 32.62, 'R$/unit', 'Fechamento de 31-jul-2026 (Status Invest / Dados de Mercado). Atualize antes de usar.', 'PxUnit', PRC),
    ('Total de ações (ON + PN)', 988880601, 'ações', 'Base acionária divulgada em agregadores de mercado, 2026.', 'AcoesTot', NUM0),
    ('Ações por unit (1 ON + 2 PN)', 3, 'ações/unit', 'Composição da unit ALUP11.', 'AcoesUnit', INT),
    ('Patrimônio líquido', 9290.0, 'R$ mn', 'Balanço mais recente divulgado (2026).', 'PL', NUM),
    ('RESULTADO 1T26 (regulatório)', None, None, None),
    ('Receita líquida consolidada 1T26', 996.8, 'R$ mn', 'Release 1T26: +16,3% a/a.', 'RecCons1T', NUM),
    ('Receita líquida transmissão 1T26', 739.0, 'R$ mn', 'Release 1T26: +16,4% a/a; TCE, reforços ELTE e TBO.', 'RecTr1T', NUM),
    ('EBITDA regulatório 1T26', 795.0, 'R$ mn', 'Release 1T26: +16% a/a, margem 79,7%.', 'EbitdaCons1T', NUM),
    ('Lucro líquido regulatório 1T26', 148.9, 'R$ mn', 'Release 1T26: +6,3% a/a.', 'Lucro1T', NUM),
    ('Lucro líquido regulatório 2025', 737.6, 'R$ mn', 'Acumulado 2025: +30,7% vs 2024.', 'Lucro25', NUM),
    ('DÍVIDA (1T26)', None, None, None),
    ('Dívida bruta', 14600.0, 'R$ mn', 'Release 1T26.', 'DivBruta', NUM),
    ('Caixa e aplicações', 5300.0, 'R$ mn', 'Release 1T26.', 'Caixa', NUM),
    ('Dívida líquida', 9300.0, 'R$ mn', 'Release 1T26; alavancagem regulatória 3,2x.', 'DivLiq0', NUM),
    ('ATIVOS E PIPELINE', None, None, None),
    ('Sistemas de transmissão em concessão', 44, 'un', '31 operacionais e 13 em implantação (operação entre 2026 e 2029).', 'NSist', INT),
    ('Capex a investir até 2029', 9100.0, 'R$ mn', 'Maior ciclo de investimentos da história da companhia.', 'CapexTot', NUM),
    ('RAP do último lote arrematado', 96.72, 'R$ mn', 'Lote 7, Leilão ANEEL 1/2026, 30 anos, energização jun/2031.', 'RapLote', NUM),
    ('Capex do último lote arrematado', 1089.0, 'R$ mn', 'Mesmo lote — a razão RAP/capex ancora o yield do pipeline.', 'CapexLote', NUM),
    ('Capacidade instalada de geração', 798.5, 'MW', 'Hidráulica, eólica e solar no Brasil, Colômbia e Peru.', 'MWinst', NUM),
]
r = 5
for row in SRC_ROWS:
    if row[1] is None:
        put(ws, r, 2, row[0], font=F_K, fill=FKEY)
        for c in (3, 4, 5):
            ws.cell(r, c).fill = FKEY
    else:
        lbl, val, un, fonte, nm, nf = row
        put(ws, r, 2, lbl, font=F_C)
        put(ws, r, 3, val, nf, F_SRC).border = BOX
        put(ws, r, 4, un, font=F_N)
        put(ws, r, 5, fonte, font=F_N).alignment = WRAP
        name(nm, SH['src'], f'$C${r}')
    r += 1

put(ws, r + 1, 2, 'Derivados diretos', font=F_K, fill=FKEY)
for c in (3, 4, 5):
    ws.cell(r + 1, c).fill = FKEY
DERIV = [
    ('Units equivalentes', f'=AcoesTot/AcoesUnit', 'units', 'Total de ações convertido em units.', 'Units', NUM0),
    ('Valor de mercado', f'=Units*PxUnit/1000000', 'R$ mn', 'Units × cotação.', 'MktCap', NUM),
    ('Firm value (mercado + dívida líquida)', '=MktCap+DivLiq0', 'R$ mn', 'Referência de entrada do DCF.', 'FirmValue', NUM),
    ('Yield de RAP sobre capex (último leilão)', '=RapLote/CapexLote', '%', 'Âncora empírica para a RAP do pipeline.', 'YieldLeilao', PCT),
    ('Receita líquida transmissão anualizada', '=RecTr1T*4', 'R$ mn', '1T26 × 4. Ignora sazonalidade e as entradas de 2026.', 'RecTrAnual', NUM),
]
r += 2
for lbl, f, un, nota, nm, nf in DERIV:
    put(ws, r, 2, lbl, font=F_C)
    put(ws, r, 3, f, nf, F_C).border = BOX
    put(ws, r, 4, un, font=F_N)
    put(ws, r, 5, nota, font=F_N).alignment = WRAP
    name(nm, SH['src'], f'$C${r}')
    r += 1

# ============================================================ Inputs (amarelo)
ws = wb.create_sheet(SH['inp'])
ws.cell(1, 1, 'Inputs — todo o modelo é dirigido por esta aba').font = F_T
ws.cell(2, 1, 'Edite apenas as células AMARELAS. Todas as demais abas recalculam sozinhas. '
              'O modelo é integralmente REAL (deflacionado por IPCA), coerente com a indexação da RAP.').font = F_N
for col, w in [('A', 3), ('B', 40), ('C', 14), ('D', 11), ('E', 70)]:
    ws.column_dimensions[col].width = w

BLOCKS = [
    ('DATA-BASE E DESCONTO', [
        ('Ano-base da avaliação', ANO0, INT, 'AnoBase',
         'Primeiro ano projetado. O fluxo desse ano é descontado a meio período (t = 0,5).'),
        ('NTN-B 10 anos (juro real)', 0.0833, PCT, 'NTNB',
         'Taxa livre de risco real. Tesouro IPCA+ 2035 (~9 anos, vértice mais líquido), IPCA+8,33% a.a. em jul/2026 — '
         'juros longos abriram no trimestre por aversão a risco externa e preocupação fiscal doméstica. Como a RAP '
         'é indexada ao IPCA, o desconto é real e comparável ao título.'),
        ('Prêmio de risco de equity (real)', 0.0450, PCT, 'ERP',
         'Prêmio exigido sobre a NTN-B. Transmissão é o ativo de menor beta do setor; 4–5 p.p. é a faixa usual.'),
        ('Custo real da dívida', 0.0550, PCT, 'KdReal',
         'Custo médio da dívida deflacionado. Alimenta a despesa de juros do fluxo ao acionista.'),
        ('Inflação de longo prazo (IPCA)', 0.0400, PCT, 'Infl',
         'Só converte o resultado real em nominal na aba Valuation. Não afeta preço-alvo nem TIR real.'),
    ]),
    ('TRANSMISSÃO — BASE OPERACIONAL', [
        ('Deduções sobre a RAP', 0.1300, PCT, 'DedTr',
         'PIS/COFINS, P&D, TFSEE e encargos. Converte RAP bruta em receita líquida — e vice-versa, que é '
         'como a RAP base é reconstruída a partir da receita divulgada.'),
        ('O&M / custos operacionais (% da RAP)', 0.1200, PCT, 'OMTr',
         'Custo gerencial de operar as linhas. A margem EBITDA implícita é conferida na aba Transmissão.'),
        ('Erosão real anual da RAP', 0.0050, PCT, 'Erosao',
         'Captura de forma reduzida as revisões tarifárias periódicas e o degrau de RAP na metade da concessão. '
         'Zere para uma RAP real perfeitamente plana.'),
        ('Início dos vencimentos de concessão', 2030, INT, 'AnoIniVenc',
         'Ano em que a RAP da base operacional começa a expirar. Ancorado em 3 concessões controladas pela Alupar '
         'com data pública: ECTE vence em 2030, EBTE em 2038, Aimorés em 2047 — a mais antiga do trio abre a janela.'),
        ('Fim dos vencimentos de concessão', 2047, INT, 'AnoFimVenc',
         'Ano em que a última concessão da base atual expira. Entre os dois anos a RAP decai linearmente até zero. '
         'Mesma âncora de 3 concessões (2030/2038/2047) — ainda uma aproximação linear de algo que é ativo a ativo, '
         'mas presa a datas reais em vez de inventada. É a hipótese que mais move o valor no longo prazo — ver Read me.'),
        ('Capex de manutenção (% da RAP)', 0.0300, PCT, 'CapexMan',
         'Investimento recorrente para manter os ativos, fora do capex de expansão.'),
    ]),
    ('TRANSMISSÃO — PIPELINE', [
        ('Eficiência de capex vs. referência ANEEL', 0.70, PCT, 'EficCapex',
         'Quanto a companhia efetivamente gasta para cada real do capex de referência do leilão. Construir '
         'abaixo da referência é de onde vem boa parte do retorno em transmissão. 0,70 = constrói ~30% abaixo, '
         'a economia de capex divulgada para o Lote 7 (ADVFN, jul/2026) — mais agressiva que a estimativa '
         'inicial de 20%. Em 100% o projeto rende apenas o yield nominal do leilão e a expansão vira destruição de valor.'),
        ('Yield de RAP sobre o capex novo', '=YieldLeilao/EficCapex', PCT, 'YieldRAP',
         'RAP por real efetivamente investido. Sai do último leilão (RAP/capex de referência do Lote 7) '
         'corrigido pela eficiência acima. Sobrescreva com um número fixo para testar outros leilões.'),
        ('Defasagem capex → RAP (anos)', 2, INT, 'Lag',
         'Quantos anos depois do desembolso a RAP começa a ser recebida. Reflete o prazo de construção.'),
        ('Capex 2026 (% do total)', 0.30, PCT1, 'CapexA', 'Distribuição dos R$ 9,1 bi entre 2026 e 2029. A soma dos quatro deve dar 100%.'),
        ('Capex 2027 (% do total)', 0.30, PCT1, 'CapexB', ''),
        ('Capex 2028 (% do total)', 0.25, PCT1, 'CapexC', ''),
        ('Capex 2029 (% do total)', 0.15, PCT1, 'CapexD', ''),
        ('Soma do cronograma de capex', '=CapexA+CapexB+CapexC+CapexD', PCT1, 'CapexSoma', 'Confira: precisa ser 100,0%.'),
        ('Prazo das concessões novas (anos)', 30, INT, 'PrazoNovo',
         'Concessões de 30 anos a partir da energização. Cada safra de capex tem vida própria.'),
        ('Último ano de novos leilões', 2045, INT, 'AnoUltLeilao',
         'Até quando a companhia segue arrematando. O horizonte vai a 2080 justamente para que a última safra '
         'de capex complete seus 30 anos dentro do modelo — assim o valor terminal é zero de verdade, e não '
         'uma perpetuidade escondendo o resultado.'),
        ('Capex recorrente pós-2029 (R$ mn/ano)', 800.0, NUM, 'CapexRec',
         'Quanto a companhia continua investindo em novos leilões depois do ciclo atual. Zerar equivale a supor '
         'que a Alupar para de crescer em 2029 e vira uma carteira em run-off — o que derruba o valor e não '
         'descreve uma companhia que arremata lote atrás de lote. É a premissa de crescimento do modelo.'),
    ]),
    ('GERAÇÃO', [
        ('Energia assegurada vendida (MW médio)', 430.0, NUM, 'MWmed',
         'Garantia física contratada, bem abaixo dos MW instalados. Premissa — a companhia não abre o número '
         'consolidado nos materiais consultados.'),
        ('Preço real da energia (R$/MWh)', 260.0, NUM, 'PxMWh',
         'Preço médio de contratação em termos reais, já considerando renovação de PPAs. Junto com o MW médio, foi calibrado para reproduzir a receita de geração implícita no 1T26 — ver a aferição na aba Consolidado.'),
        ('Deduções sobre receita de geração', 0.1500, PCT, 'DedGe', 'PIS/COFINS e encargos setoriais.'),
        ('Opex de geração (% da receita líquida)', 0.2500, PCT, 'OMGe', 'O&M, seguros e encargos de uso da rede.'),
        ('Receita de comercialização (R$ mn/ano)', 200.0, NUM, 'ComercRec',
         'A Alup, comercializadora do grupo. Entra separada porque gira volume com margem fina — jogá-la dentro '
         'da energia assegurada inflaria o MW médio e, com ele, a margem da geração.'),
        ('Margem EBITDA da comercialização', 0.0800, PCT, 'ComercMg',
         'Margem típica de trading de energia. Baixa por natureza.'),
        ('Ano final das outorgas de geração', 2045, INT, 'AnoFimGer',
         'Fim médio das outorgas. Depois disso a geração sai do fluxo — renovação vira upside não modelado.'),
    ]),
    ('HOLDING, DÍVIDA E IMPOSTOS', [
        ('Custos de holding (R$ mn/ano, real)', 120.0, NUM, 'Holding', 'Estrutura corporativa, fora das SPEs.'),
        ('Carga tributária efetiva (% do EBITDA)', 0.1000, PCT, 'TaxEff',
         'IR/CSLL de caixa. Muito abaixo dos 34% nominais: as SPEs de transmissão apuram em lucro presumido, '
         'em que o imposto sai como ~3–4% da receita bruta, e boa parte do portfólio tem redução de IR por '
         'SUDAM/SUDENE. Contra um EBITDA da ordem de R$ 3 bi, isso dá algo entre 5% e 10%. Colocar os 20% que '
         'pareciam prudentes derruba o valor em quase 40% e faz cada novo leilão destruir valor — o que não '
         'descreve esta companhia.'),
        ('Minoritários (% do fluxo operacional)', 0.1200, PCT, 'Minor',
         'A Alupar consolida SPEs com sócios. Esta fatia do fluxo não pertence ao acionista de ALUP11.'),
        ('% do capex financiado com dívida', 0.6000, PCT, 'PctDivida', 'O restante vem do caixa gerado.'),
        ('Prazo médio de amortização (anos)', 10, INT, 'PrazoDiv', 'Amortização linear sobre o saldo do ano anterior.'),
        ('Valor terminal ao acionista (R$ mn, real)', 0.0, NUM, 'VT',
         'Zero por construção: as concessões expiram dentro do horizonte. Preencha para precificar renovação '
         'ou novos leilões após 2029 — é aqui que entra a opcionalidade de crescimento perpétuo.'),
    ]),
]

r = 4
INPUT_ROW = {}
for title, items in BLOCKS:
    put(ws, r, 2, title, font=F_K, fill=FKEY)
    for c in (3, 4, 5):
        ws.cell(r, c).fill = FKEY
    r += 1
    for lbl, val, nf, nm, nota in items:
        put(ws, r, 2, lbl, font=F_C)
        is_formula = isinstance(val, str) and val.startswith('=')
        cell = put(ws, r, 3, val, nf, F_C if is_formula else F_IN, None if is_formula else FIN)
        cell.border = BOX
        put(ws, r, 5, nota, font=F_N).alignment = WRAP
        name(nm, SH['inp'], f'$C${r}')
        INPUT_ROW[nm] = r
        r += 1
    r += 1

put(ws, r, 2, 'Ke real (NTN-B + prêmio)', font=F_K)
put(ws, r, 3, '=NTNB+ERP', PCT, F_K, FKEY).border = BOX
name('Ke', SH['inp'], f'$C${r}')
put(ws, r, 5, 'Taxa usada para descontar todo o fluxo ao acionista. Real, porque o fluxo é real.', font=F_N)
r += 1
put(ws, r, 2, 'Ke nominal implícito', font=F_C)
put(ws, r, 3, '=(1+Ke)*(1+Infl)-1', PCT, F_C).border = BOX
put(ws, r, 5, 'Apenas informativo — nada no modelo usa esta taxa.', font=F_N)

# ============================================================ Transmissão
ws = wb.create_sheet(SH['tr'])
ws.cell(1, 1, 'Transmissão — RAP real, base operacional e pipeline (R$ mn, moeda de hoje)').font = F_T
ws.cell(2, 1, 'A RAP da base operacional é reconstruída a partir da receita líquida divulgada no 1T26; '
              'a RAP do pipeline é o capex acumulado (com defasagem) multiplicado pelo yield de leilão.').font = F_N
COLS = ['Ano', 'Capex expansão', 'Capex acumulado', 'Capex de safras vivas', 'RAP pipeline',
        'RAP base operacional', 'Fator de vencimento', 'RAP total', 'Deduções', 'Receita líquida',
        'O&M', 'EBITDA', 'Margem EBITDA', 'Capex manutenção']
head(ws, 4, COLS, [8] + [13] * 3 + [12, 14, 12, 11, 11, 12, 10, 11, 11, 12])
ws.freeze_panes = 'B5'

put(ws, 5, 2, '=CapexTot*CapexA', NUM, F_C)
put(ws, 6, 2, '=CapexTot*CapexB', NUM, F_C)
put(ws, 7, 2, '=CapexTot*CapexC', NUM, F_C)
put(ws, 8, 2, '=CapexTot*CapexD', NUM, F_C)

RAPBASE = 'RecTrAnual/(1-DedTr)'          # RAP bruta implícita na receita líquida divulgada
for i, ano in enumerate(ANOS):
    r = 5 + i
    put(ws, r, 1, ano, INT, F_K)
    if i >= 4:
        put(ws, r, 2, f'=IF(A{r}<=AnoUltLeilao,CapexRec,0)', NUM, F_C)
    put(ws, r, 3, f'=B{r}' if i == 0 else f'=C{r-1}+B{r}', NUM, F_C)
    # safras vivas no ano t: desembolsadas até t-Lag e ainda dentro dos PrazoNovo anos de concessão
    put(ws, r, 4, f'=SUMPRODUCT(($A$5:$A${5+N-1}<=A{r}-Lag)*($A$5:$A${5+N-1}>A{r}-Lag-PrazoNovo)'
                  f'*$B$5:$B${5+N-1})', NUM, F_C)
    put(ws, r, 5, f'=D{r}*YieldRAP*(1-Erosao)^(A{r}-AnoBase)', NUM, F_C)
    put(ws, r, 6, f'={RAPBASE}*(1-Erosao)^(A{r}-AnoBase)', NUM, F_C)
    put(ws, r, 7, f'=IF(A{r}<AnoIniVenc,1,MAX(0,1-(A{r}-AnoIniVenc+1)/(AnoFimVenc-AnoIniVenc+1)))', PCT1, F_C)
    put(ws, r, 8, f'=F{r}*G{r}+E{r}', NUM, F_C)
    put(ws, r, 9, f'=-H{r}*DedTr', NUM, F_C)
    put(ws, r, 10, f'=H{r}+I{r}', NUM, F_C)
    put(ws, r, 11, f'=-H{r}*OMTr', NUM, F_C)
    put(ws, r, 12, f'=J{r}+K{r}', NUM, F_K)
    put(ws, r, 13, f'=IFERROR(L{r}/J{r},"")', PCT1, F_C)
    put(ws, r, 14, f'=-H{r}*CapexMan', NUM, F_C)

TR_LAST = 5 + N - 1
ch = LineChart()
ch.title = 'RAP real da transmissão (R$ mn)'
ch.height, ch.width = 8, 18
ch.y_axis.title, ch.x_axis.title = 'R$ mn (moeda de hoje)', None
for col, lab in [(6, 'Base operacional'), (5, 'Pipeline'), (8, 'RAP total')]:
    ref = Reference(ws, min_col=col, min_row=4, max_row=TR_LAST)
    ch.add_data(ref, titles_from_data=True)
ch.set_categories(Reference(ws, min_col=1, min_row=5, max_row=TR_LAST))
ws.add_chart(ch, f'P4')

# ============================================================ Geração
ws = wb.create_sheet(SH['ge'])
ws.cell(1, 1, 'Geração — 798,5 MW instalados, fluxo contratado em termos reais (R$ mn)').font = F_T
ws.cell(2, 1, 'Bloco deliberadamente simples: energia assegurada contratada a preço real constante até o fim das '
              'outorgas. É a parte do modelo com menos dado público por trás.').font = F_N
head(ws, 4, ['Ano', 'Em outorga', 'Energia (MW médio)', 'Receita bruta', 'Deduções',
             'Receita líquida', 'Opex', 'EBITDA geração', 'Receita de comercialização',
             'EBITDA comercialização', 'Receita líquida total', 'EBITDA total'],
     [8, 11, 15, 13, 11, 13, 11, 14, 15, 15, 14, 13])
ws.freeze_panes = 'B5'
for i, ano in enumerate(ANOS):
    r = 5 + i
    put(ws, r, 1, ano, INT, F_K)
    put(ws, r, 2, f'=IF(A{r}<=AnoFimGer,1,0)', INT, F_C)
    put(ws, r, 3, f'=MWmed*B{r}', NUM, F_C)
    put(ws, r, 4, f'=C{r}*8760*PxMWh/1000000', NUM, F_C)
    put(ws, r, 5, f'=-D{r}*DedGe', NUM, F_C)
    put(ws, r, 6, f'=D{r}+E{r}', NUM, F_C)
    put(ws, r, 7, f'=-F{r}*OMGe', NUM, F_C)
    put(ws, r, 8, f'=F{r}+G{r}', NUM, F_C)
    put(ws, r, 9, f'=ComercRec*B{r}', NUM, F_C)
    put(ws, r, 10, f'=I{r}*ComercMg', NUM, F_C)
    put(ws, r, 11, f'=F{r}+I{r}', NUM, F_C)
    put(ws, r, 12, f'=H{r}+J{r}', NUM, F_K)

# ============================================================ Consolidado (FCFE)
ws = wb.create_sheet(SH['cons'])
ws.cell(1, 1, 'Consolidado — fluxo de caixa ao acionista (FCFE) em termos reais (R$ mn)').font = F_T
ws.cell(2, 1, 'EBITDA das duas pontas menos holding, impostos, capex, minoritários e serviço da dívida. '
              'A dívida rola: capta uma fração do capex de expansão e amortiza linearmente.').font = F_N
head(ws, 4, ['Ano', 'EBITDA transmissão', 'EBITDA geração + comerc.', 'Holding', 'EBITDA consolidado',
             'Impostos efetivos', 'Capex expansão', 'Capex manutenção', 'Minoritários',
             'Fluxo antes da dívida', 'Dívida — saldo inicial', 'Captação', 'Amortização',
             'Juros reais', 'Dívida — saldo final', 'Alavancagem (DL/EBITDA)', 'FCFE'],
     [8] + [13] * 4 + [12, 12, 12, 12, 14, 13, 11, 12, 11, 13, 13, 12])
ws.freeze_panes = 'B5'
for i, ano in enumerate(ANOS):
    r, t = 5 + i, 5 + i
    put(ws, r, 1, ano, INT, F_K)
    put(ws, r, 2, f"='{SH['tr']}'!L{t}", NUM, F_C)
    put(ws, r, 3, f"='{SH['ge']}'!L{t}", NUM, F_C)
    put(ws, r, 4, '=-Holding', NUM, F_C)
    put(ws, r, 5, f'=B{r}+C{r}+D{r}', NUM, F_K)
    put(ws, r, 6, f'=-MAX(0,E{r})*TaxEff', NUM, F_C)
    put(ws, r, 7, f"=-'{SH['tr']}'!B{t}", NUM, F_C)
    put(ws, r, 8, f"='{SH['tr']}'!N{t}", NUM, F_C)
    put(ws, r, 9, f'=-MAX(0,E{r}+F{r})*Minor', NUM, F_C)
    put(ws, r, 10, f'=SUM(E{r}:I{r})', NUM, F_K)
    put(ws, r, 11, '=DivLiq0' if i == 0 else f'=O{r-1}', NUM, F_C)
    put(ws, r, 12, f'=-G{r}*PctDivida', NUM, F_C)
    put(ws, r, 13, f'=-K{r}/PrazoDiv', NUM, F_C)
    put(ws, r, 14, f'=-K{r}*KdReal', NUM, F_C)
    put(ws, r, 15, f'=K{r}+L{r}+M{r}', NUM, F_C)
    put(ws, r, 16, f'=IFERROR(O{r}/E{r},"")', '0.00"x"', F_C)
    put(ws, r, 17, f'=J{r}+L{r}+M{r}+N{r}', NUM, F_K)

CONS_LAST = 5 + N - 1

rc = CONS_LAST + 3
put(ws, rc, 1, 'AFERIÇÃO CONTRA O DIVULGADO — se estas linhas destoarem, recalibre os Inputs antes de ler o valuation',
    font=F_K, fill=FKEY)
for c in range(2, 6):
    ws.cell(rc, c).fill = FKEY
CHECKS = [
    ('EBITDA consolidado 2026 do modelo', '=E5', NUM, ''),
    ('EBITDA regulatório 1T26 anualizado', '=EbitdaCons1T*4', NUM, 'Release 1T26 × 4.'),
    ('Desvio', f'=IFERROR(B{rc+1}/B{rc+2}-1,"")', PCT1, 'Tolerável até ~±10%: o run-rate do 1T26 já embute '
                                                        'entradas em operação que o modelo distribui ao longo do ano.'),
    ('Receita de geração + comercialização 2026', f"='{SH['ge']}'!K5", NUM, ''),
    ('Receita de geração implícita no 1T26', '=(RecCons1T-RecTr1T)*4', NUM,
     'Receita consolidada menos transmissão, anualizada. Foi contra esta linha que MW médio e preço foram calibrados.'),
    ('Desvio', f'=IFERROR(B{rc+4}/B{rc+5}-1,"")', PCT1, ''),
    ('Alavancagem máxima até 2035', '=MAX(P5:P14)', '0.00"x"',
     'A janela para em 2035 de propósito: depois que as concessões começam a expirar o EBITDA cai e o índice '
     'dispara sem que isso signifique estresse de crédito — é aritmética de run-off, não alavancagem.'),
    ('Ano do pico de alavancagem', f'=INDEX(A5:A14,MATCH(B{rc+7},P5:P14,0))', INT, ''),
    ('Guidance da companhia', 3.95, '0.00"x"', 'Pico de 3,9–4,0x em 2028, segundo a diretoria de RI. '
                                               'É a melhor aferição disponível do cronograma de capex.'),
]
for k, (lbl, f, nf, nota) in enumerate(CHECKS, 1):
    put(ws, rc + k, 1, lbl, font=F_C)
    put(ws, rc + k, 2, f, nf, F_K if 'Desvio' not in lbl else F_C).border = BOX
    put(ws, rc + k, 4, nota, font=F_N)
ws.cell(rc + 9, 2).font = F_SRC

# ============================================================ Valuation
ws = wb.create_sheet(SH['val'])
ws.cell(1, 1, 'Valuation — DCF do fluxo ao acionista e TIR real implícita').font = F_T
ws.cell(2, 1, 'Fluxo real descontado ao Ke real, convenção de meio de ano. A TIR real implícita compara-se '
              'diretamente com a NTN-B, porque ambas são taxas reais.').font = F_N
head(ws, 4, ['Ano', 't (meio de ano)', 'FCFE real', 'Fator de desconto', 'Valor presente',
             'VP acumulado', 'Fluxo p/ TIR'], [8, 14, 13, 14, 13, 13, 13])
ws.freeze_panes = 'B5'
put(ws, 5, 1, 'Hoje', font=F_K)
put(ws, 5, 7, '=-MktCap', NUM, F_K)
for i, ano in enumerate(ANOS):
    r = 6 + i
    put(ws, r, 1, ano, INT, F_K)
    put(ws, r, 2, f'=A{r}-AnoBase+0.5', '0.0', F_C)
    put(ws, r, 3, f"='{SH['cons']}'!Q{5+i}", NUM, F_C)
    put(ws, r, 4, f'=1/(1+Ke)^B{r}', '0.0000', F_C)
    put(ws, r, 5, f'=C{r}*D{r}', NUM, F_C)
    put(ws, r, 6, f'=E{r}' if i == 0 else f'=F{r-1}+E{r}', NUM, F_C)
    put(ws, r, 7, f'=C{r}' + (f'+VT' if i == N - 1 else ''), NUM, F_C)
VAL_LAST = 6 + N - 1

r0 = VAL_LAST + 3
put(ws, r0, 1, 'RESULTADO', font=F_K, fill=FKEY)
for c in range(2, 5):
    ws.cell(r0, c).fill = FKEY
RES = [
    ('VP do fluxo explícito', f'=SUM(E6:E{VAL_LAST})', NUM, 'R$ mn'),
    ('VP do valor terminal', f'=VT/(1+Ke)^B{VAL_LAST}', NUM, 'R$ mn'),
    ('Valor do equity', f'=B{r0+1}+B{r0+2}', NUM, 'R$ mn'),
    ('Valor de mercado atual', '=MktCap', NUM, 'R$ mn'),
    ('Preço-alvo por unit', f'=B{r0+3}*1000000/Units', PRC, 'R$/unit'),
    ('Cotação atual', '=PxUnit', PRC, 'R$/unit'),
    ('Upside / (downside)', f'=B{r0+5}/B{r0+6}-1', PCT, ''),
    ('TIR real implícita (no preço atual)', f'=IRR(G5:G{VAL_LAST},Ke)', PCT, 'a.a. real'),
    ('TIR nominal equivalente', f'=(1+B{r0+8})*(1+Infl)-1', PCT, 'a.a.'),
    ('NTN-B 10 anos', '=NTNB', PCT, 'a.a. real'),
    ('Prêmio sobre a NTN-B', f'=B{r0+8}-B{r0+10}', PCT, 'p.p.'),
    ('Gap a explicar (mercado − DCF)', f'=B{r0+4}-B{r0+3}', NUM, 'R$ mn'),
    ('Gap por unit', f'=B{r0+12}*1000000/Units', PRC, 'R$/unit'),
]
for k, (lbl, f, nf, un) in enumerate(RES, 1):
    put(ws, r0 + k, 1, lbl, font=F_C)
    put(ws, r0 + k, 2, f, nf, F_K).border = BOX
    put(ws, r0 + k, 3, un, font=F_N)
name('Equity', SH['val'], f'$B${r0+3}')
name('Alvo', SH['val'], f'$B${r0+5}')
name('TIRreal', SH['val'], f'$B${r0+8}')
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 15

rs = r0 + len(RES) + 2
put(ws, rs, 1, 'ABERTURA POR SEGMENTO (VP do EBITDA após impostos, R$ mn)', font=F_K, fill=FKEY)
for c in range(2, 5):
    ws.cell(rs, c).fill = FKEY
SOTP = [
    ('Transmissão', f"=SUMPRODUCT('{SH['cons']}'!B5:B{CONS_LAST}*(1-TaxEff),D6:D{VAL_LAST})"),
    ('Geração e comercialização', f"=SUMPRODUCT('{SH['cons']}'!C5:C{CONS_LAST}*(1-TaxEff),D6:D{VAL_LAST})"),
    ('Holding', f"=SUMPRODUCT('{SH['cons']}'!D5:D{CONS_LAST},D6:D{VAL_LAST})"),
    ('(-) Dívida líquida atual', '=-DivLiq0'),
    ('(-) Minoritários e capex, líquido', f'=Equity-B{rs+1}-B{rs+2}-B{rs+3}-B{rs+4}'),
    ('Valor do equity', f'=SUM(B{rs+1}:B{rs+5})'),
]
for k, (lbl, f) in enumerate(SOTP, 1):
    put(ws, rs + k, 1, lbl, font=F_C if k < 6 else F_K)
    put(ws, rs + k, 2, f, NUM, F_C if k < 6 else F_K).border = BOX
put(ws, rs + 7, 1, 'A abertura é indicativa: descontar EBITDA por segmento não é o mesmo que alocar capex e '
                   'dívida a cada um. A linha residual fecha a conta contra o equity do DCF.', font=F_N)

# ============================================================ Sensibilidade
ws = wb.create_sheet(SH['sens'])
ws.cell(1, 1, 'Sensibilidade').font = F_T
ws.cell(2, 1, 'Cada célula reavalia o DCF inteiro pela fórmula — não depende de tabela de dados do Excel, '
              'então os números aparecem corretos em qualquer leitor de planilha.').font = F_N
ws.column_dimensions['A'].width = 26
for c in range(2, 9):
    ws.column_dimensions[L(c)].width = 12

FLOW = f"'{SH['val']}'!$C$6:$C${VAL_LAST}"
TCOL = f"'{SH['val']}'!$B$6:$B${VAL_LAST}"

put(ws, 4, 1, 'Preço-alvo por unit (R$)', font=F_K, fill=FKEY)
put(ws, 5, 1, 'Ke real  ↓   /   Choque no fluxo  →', font=F_N)
SHOCKS = [-0.15, -0.075, 0.0, 0.075, 0.15]
KES = [-0.02, -0.01, 0.0, 0.01, 0.02]
for j, s in enumerate(SHOCKS, 2):
    put(ws, 5, j, s, PCT1, F_H, FH).alignment = Alignment(horizontal='center')
for i, dk in enumerate(KES, 6):
    put(ws, i, 1, f'=Ke+{dk}', PCT, F_H, FH)
    for j, s in enumerate(SHOCKS, 2):
        put(ws, i, j,
            f'=(SUMPRODUCT({FLOW}*(1+{s}),1/(1+$A{i})^{TCOL})+VT/(1+$A{i})^{N-0.5})*1000000/Units',
            PRC, F_C).border = BOX

r2 = 13
put(ws, r2, 1, 'Upside vs. cotação atual', font=F_K, fill=FKEY)
put(ws, r2 + 1, 1, 'Ke real  ↓   /   Choque no fluxo  →', font=F_N)
for j, s in enumerate(SHOCKS, 2):
    put(ws, r2 + 1, j, s, PCT1, F_H, FH).alignment = Alignment(horizontal='center')
for i in range(5):
    rr = r2 + 2 + i
    put(ws, rr, 1, f'=A{6+i}', PCT, F_H, FH)
    for j in range(2, 7):
        put(ws, rr, j, f'={L(j)}{6+i}/PxUnit-1', PCT, F_C).border = BOX

r3 = 21
put(ws, r3, 1, 'TIR real implícita por preço de entrada', font=F_K, fill=FKEY)
put(ws, r3 + 1, 1, 'Preço da unit (R$)', font=F_H, fill=FH)
put(ws, r3 + 2, 1, 'TIR real implícita', font=F_H, fill=FH)
put(ws, r3 + 3, 1, 'Prêmio sobre a NTN-B', font=F_H, fill=FH)
rf = r3 + 6                                   # bloco auxiliar de fluxos, um por preço
put(ws, rf - 1, 1, 'Fluxos usados acima (auxiliar) — ano 0 = desembolso do investidor', font=F_N)
put(ws, rf, 1, 'Ano', font=F_H, fill=FH)
for i in range(N + 1):
    put(ws, rf + 1 + i, 1, 'Hoje' if i == 0 else ANOS[i - 1], INT if i else None, F_C)
for j, mult in enumerate([0.8, 0.9, 1.0, 1.1, 1.2, 1.3], 2):
    c = L(j)
    put(ws, r3 + 1, j, f'=PxUnit*{mult}', PRC, F_C).border = BOX
    put(ws, r3 + 2, j, f'=IRR({c}{rf+1}:{c}{rf+N+1},0.03)', PCT, F_K).border = BOX
    put(ws, r3 + 3, j, f'={c}{r3+2}-NTNB', PCT, F_C).border = BOX
    put(ws, rf, j, f'={c}{r3+1}', PRC, F_H, FH)
    put(ws, rf + 1, j, f'=-{c}{r3+1}*Units/1000000', NUM, F_C)
    for i in range(N):
        put(ws, rf + 2 + i, j, f"='{SH['val']}'!$G${6+i}", NUM, F_C)
ws.row_dimensions[rf - 1].height = 14

# ============================================================ Read me
ws = wb.create_sheet(SH['read'], 0)
ws.cell(1, 1, 'Alupar (ALUP11) — modelo de valuation').font = F_T
ws.column_dimensions['A'].width = 118
TEXT = [
    ('k', 'O que este arquivo é'),
    ('p', 'Um DCF do fluxo de caixa ao acionista da Alupar, montado em termos REAIS (moeda de hoje). '
          'A escolha não é estética: a RAP da transmissão é corrigida por IPCA, então projetar em real e '
          'descontar a uma taxa real evita embutir inflação duas vezes — e deixa a TIR resultante diretamente '
          'comparável com a NTN-B.'),
    ('p', 'O workbook inteiro é dirigido por fórmulas. Mexer numa célula amarela de Inputs recalcula tudo, '
          'inclusive as sensibilidades. Não existe número calculado gravado como valor.'),
    ('k', 'Código de cores'),
    ('p', 'AZUL sobre amarelo, na aba Inputs — premissa editável, é sua para calibrar.\n'
          'AZUL na aba Dados de origem — número público coletado, com a fonte ao lado.\n'
          'PRETO — fórmula. Não digite por cima.'),
    ('k', 'Como o modelo é construído'),
    ('p', '1. TRANSMISSÃO. A RAP bruta da base operacional é reconstruída a partir da receita líquida '
          'regulatória de transmissão do 1T26 anualizada, revertida pela alíquota de deduções. O pipeline é '
          'tratado em bloco: o capex de R$ 9,1 bi é distribuído entre 2026 e 2029, e cada real desembolsado '
          'passa a gerar RAP dois anos depois, ao yield do último leilão (RAP/capex do Lote 7 = 8,9%). A base '
          'atual expira linearmente entre 2030 e 2047 (âncoras: ECTE, EBTE e Aimorés); o pipeline roda 30 anos a '
          'partir da energização.'),
    ('p', '2. GERAÇÃO. Energia assegurada contratada a preço real constante até o fim das outorgas. É a parte '
          'mais frágil do modelo — a companhia não abre o MW médio consolidado nos materiais consultados, então '
          'o número é premissa, não dado.'),
    ('p', '3. CONSOLIDADO. EBITDA das duas pontas menos holding, imposto efetivo, capex e minoritários. A dívida '
          'rola de verdade: capta uma fração do capex de expansão, amortiza linearmente e paga juros reais sobre '
          'o saldo. A coluna de alavancagem serve de teste de sanidade — a companhia projeta pico de 3,9–4,0x '
          'em 2028, e o modelo deve chegar perto disso com as premissas padrão.'),
    ('p', '4. VALUATION. Desconto a meio de ano ao Ke real (NTN-B + prêmio). O preço-alvo sai do equity dividido '
          'pelas units equivalentes; a TIR real implícita resolve a taxa que iguala o fluxo ao valor de mercado '
          'de hoje.'),
    ('k', 'Onde o modelo é fraco — leia antes de citar qualquer número'),
    ('p', '• O CRONOGRAMA DE VENCIMENTOS É APROXIMADO, AGORA COM ÂNCORAS REAIS. A base operacional são 31 '
          'sistemas com datas de outorga diferentes, e a lista completa não está disponível publicamente. O '
          'decaimento linear entre 2030 e 2047 usa as únicas três datas de concessão da Alupar que encontramos '
          'publicadas — ECTE (2030), EBTE (2038) e Aimorés (2047) — em vez de um intervalo inventado. Ainda é '
          'uma média de três pontos aplicada a 31 ativos, não um cronograma ativo a ativo, e seguem sendo a '
          'premissa que mais move o valor no longo prazo. Substitua-a assim que tiver a lista completa da ANEEL.'),
    ('p', '• O PIPELINE É TRATADO EM BLOCO. Treze projetos com capex, RAP e datas próprias viram um único '
          'agregado ao yield do último leilão. Projetos antigos foram arrematados em condições diferentes, e a '
          'RAP efetiva de cada um não é a do Lote 7.'),
    ('p', '• A GERAÇÃO É UMA CAIXA-PRETA. MW médio, preço e opex são todos premissa. Se a geração importa para '
          'a sua tese, este bloco precisa ser refeito com os PPAs reais.'),
    ('p', '• O IMPOSTO É UMA ALÍQUOTA ÚNICA SOBRE O EBITDA. Não há depreciação, ágio, JCP nem prejuízo fiscal '
          'acumulado. Para uma holding com dezenas de SPEs em regimes distintos, é uma simplificação forte — e '
          'é também o input mais perigoso do arquivo, pelo motivo descrito no item seguinte.'),
    ('k', 'O que a montagem do modelo revelou'),
    ('p', 'Vale registrar, porque muda como se lê o resultado.'),
    ('p', '• A ALÍQUOTA EFETIVA DOMINA TUDO. A primeira versão usava 20% do EBITDA, que parecia prudente. Com '
          'ela o preço-alvo caía cerca de 40% E, pior, cada real de capex novo passava a destruir valor: '
          'aumentar o programa de investimentos reduzia o preço-alvo. Isso não descreve uma companhia que '
          'compõe valor arrematando lote atrás de lote há décadas — era sinal de erro, não de conservadorismo. '
          'A alíquota correta é bem menor: as SPEs de transmissão apuram em lucro presumido (imposto ≈ 3–4% da '
          'receita bruta) e boa parte do portfólio tem redução de IR por SUDAM/SUDENE. Daí o padrão de 10%. Se '
          'você mexer neste input, confira sempre se a expansão continua criando valor — é o melhor teste de '
          'sanidade do arquivo.'),
    ('p', '• A ECONOMIA DE CAPEX DO LOTE 7 FOI REVISADA PARA CIMA. A versão anterior assumia que a Alupar '
          'constrói a 80% do capex de referência da ANEEL (20% de economia), uma estimativa de trabalho. A '
          'imprensa especializada (ADVFN, jul/2026) noticiou uma economia de ~30% para esse lote especificamente, '
          'então o input passou para 70%. Isso melhora o retorno por real investido no pipeline e empurra o valor '
          'na direção contrária à da NTN-B — mas é um efeito de segunda ordem perto do impacto da taxa de desconto.'),
    ('p', '• A NTN-B ESTAVA DESATUALIZADA, E ISSO MUDA O RESULTADO MAIS DO QUE QUALQUER OUTRA CORREÇÃO. A versão '
          'anterior usava 7,00% real (dado de meados de julho). Em 31-jul-2026 o vértice mais líquido do Tesouro '
          'IPCA+ (2035, ~9 anos) pagava 8,33% real — juros longos abriram no trimestre por tensão geopolítica e '
          'preocupação fiscal doméstica. Como o fluxo roda até 2080, o efeito de 133 p.p.-base a mais de desconto '
          'composto por décadas é enorme: o Ke real vai de 11,5% para 12,83%, e o preço-alvo cai de ~R$ 21 para '
          'R$ 11,62/unit — quase pela metade — sem que uma única premissa operacional tenha mudado.'),
    ('p', '• O MODELO NÃO REPRODUZ O PREÇO DE MERCADO, e não se deve forçá-lo a isso. Com as premissas padrão e a '
          'NTN-B atualizada, o DCF dá cerca de R$ 11,62/unit contra uma cotação de R$ 32,62 (31-jul-2026) — um '
          'gap de ~R$ 21/unit, o dobro do que era antes de corrigir a taxa. Mais revelador: a TIR real implícita '
          'no preço atual cai para 4,75% a.a., ABAIXO dos 8,33% da própria NTN-B — quem compra ALUP11 hoje, pelo '
          'fluxo deste modelo, aceita um retorno real menor que o do título público que deveria ser seu piso. As '
          'linhas "Gap a explicar" na aba Valuation medem essa diferença. Ela é a pergunta de investimento, não '
          'um defeito a ser calibrado: o mercado está pagando por renovação de concessões, leilões além de 2045, '
          'sinergias ou um custo de capital menor do que os 12,83% reais assumidos aqui — ou o modelo real ainda '
          'está com viés pessimista demais na base operacional. A leitura mais robusta do arquivo é a TIR real '
          'implícita, não o preço-alvo — ela não depende de escolher um Ke.'),
    ('p', '• O MODELO FOI AFERIDO CONTRA TRÊS ÂNCORAS DIVULGADAS, e o resultado está na parte de baixo da aba '
          'Consolidado: EBITDA de 2026 a 3,5% do run-rate do 1T26, receita de geração e comercialização a 0,1% '
          'da implícita no trimestre, e pico de alavancagem de 3,5x em 2027 contra guidance de 3,9–4,0x em '
          '2028. Recalibre os Inputs se qualquer uma dessas linhas destoar.'),
    ('p', '• OS DADOS DE ORIGEM VÊM DE AGREGADORES E DA IMPRENSA, não dos arquivos da companhia. O acesso direto '
          'ao site de RI e aos releases foi bloqueado pela rede na montagem. Confira a aba Dados de origem '
          'contra o release oficial antes de usar isto para qualquer coisa que importe.'),
    ('k', 'Uso'),
    ('p', 'Material de estudo. Não é recomendação de investimento, e os números não foram auditados contra as '
          'demonstrações financeiras da companhia. O workbook foi recalculado de ponta a ponta e não contém '
          'nenhuma célula de erro, mas isso atesta a aritmética, não as premissas.'),
]
r = 3
for kind, txt in TEXT:
    c = put(ws, r, 1, txt, font=F_K if kind == 'k' else F_C)
    if kind == 'k':
        c.fill = FKEY
    else:
        c.alignment = WRAP
        ws.row_dimensions[r].height = 15 * (1 + txt.count('\n') + len(txt) // 105)
    r += 2 if kind == 'k' else 1

wb.active = 0
wb.save('alupar_valuation_model.xlsx')
print('ok — alupar_valuation_model.xlsx')
