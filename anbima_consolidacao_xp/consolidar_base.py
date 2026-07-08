"""Consolida o relatorio mensal da ANBIMA (Dados ambima.xlsx) com a base
historica propria (Historico.xlsx), gerando a base consolidada (Base XP.xlsx).

O script le os ultimos N meses (padrao: 12) do relatorio da ANBIMA e faz
*upsert* deles sobre o historico: meses ja existentes tem os valores
substituidos pela vintage mais recente da ANBIMA e meses novos sao
acrescentados, preservando todo o restante do historico intacto.

Abas consolidadas (nome no Historico  <-  fonte no relatorio ANBIMA):
  - 'PL Const. Por Classe'  <-  'Pag. 3 - PL Const. por Categ.'
  - 'Cap. Liq. Por Classe'  <-  'Pag. 8 - Cap. Liq. por Classe'
        (linhas de total anual dos anos afetados tambem sao atualizadas)
  - 'PL por Tipo'           <-  'Pag. 5 - PL por Tipo'   (bloco Acoes, transposto)
  - 'Cap. Liq. Tipo'        <-  'Pag. 9 - Cap. Liq. por Tipo' (bloco Acoes, transposto)

Uso:
    python consolidar_base.py
    python consolidar_base.py --meses 12 ^
        --anbima "caminho\\Dados ambima.xlsx" ^
        --historico "caminho\\Historico.xlsx" ^
        --saida "caminho\\Base XP.xlsx"
"""
import argparse
import copy
import unicodedata
from datetime import date, datetime

import openpyxl

CAMINHO_ANBIMA = r"\\xpdocs\Research\Equities\Estrategia\Reports\Fluxo investidores na Bolsa\Banco de dados\Anbima Automation\input\Dados ambima.xlsx"
CAMINHO_HISTORICO = r"\\xpdocs\Research\Equities\Estrategia\Reports\Fluxo investidores na Bolsa\Banco de dados\Anbima Automation\input\Histórico.xlsx"
CAMINHO_SAIDA = r"\\xpdocs\Research\Equities\Estrategia\Reports\Fluxo investidores na Bolsa\Banco de dados\Anbima Automation\input\Base XP.xlsx"

MESES_PT = {"jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
            "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12}


def sem_acento(texto: str) -> str:
    return unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode()


def aba(workbook, nome_alvo):
    """Localiza uma aba por nome ignorando acentos e espacos nas pontas."""
    alvo = sem_acento(nome_alvo).strip().lower()
    for nome in workbook.sheetnames:
        if sem_acento(nome).strip().lower() == alvo:
            return workbook[nome]
    raise KeyError(f"Aba '{nome_alvo}' nao encontrada. Abas: {workbook.sheetnames}")


def parse_mes(valor):
    """Converte um rotulo de periodo em (ano, mes). Aceita datetime, inteiros
    AAAAMM e strings tipo 'mai-26' ou 'mai/26'. Retorna None se nao for mes."""
    if isinstance(valor, (datetime, date)):
        return (valor.year, valor.month)
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        inteiro = int(valor)
        if 190001 <= inteiro <= 299912 and 1 <= inteiro % 100 <= 12:
            return (inteiro // 100, inteiro % 100)
        return None
    if isinstance(valor, str):
        texto = sem_acento(valor).strip().lower().replace("/", "-")
        partes = texto.split("-")
        if len(partes) == 2 and partes[0] in MESES_PT and partes[1].isdigit():
            ano = int(partes[1])
            ano += 2000 if ano < 100 else 0
            return (ano, MESES_PT[partes[0]])
    return None


# ---------------------------------------------------------------------------
# Leitura do relatorio ANBIMA
# ---------------------------------------------------------------------------

def ultimos_meses_pag_classe(ws, colunas, n_meses):
    """Le uma pagina 'por classe/categoria' (meses nas linhas). `colunas` mapeia
    nome -> indice da coluna na planilha. Retorna {(ano, mes): {nome: valor}}
    apenas para os ultimos n_meses."""
    linhas_mes = []
    for linha in ws.iter_rows(values_only=True):
        chave = parse_mes(linha[0])
        if chave is not None:
            linhas_mes.append((chave, linha))
    dados = {}
    for chave, linha in sorted(linhas_mes, key=lambda item: item[0])[-n_meses:]:
        dados[chave] = {nome: linha[indice] for nome, indice in colunas.items()}
    return dados


def ultimos_meses_pag_tipo(ws, rotulos, n_meses):
    """Le uma pagina 'por tipo' (meses nas colunas, tipos nas linhas) e extrai o
    bloco de Acoes. `rotulos` sao os nomes esperados na coluna B; o rotulo
    'Acoes' (total da categoria) so casa com a primeira ocorrencia, que antecede
    os subtipos. Retorna {(ano, mes): {rotulo: valor}}."""
    linhas = list(ws.iter_rows(values_only=True))
    cabecalho = next(l for l in linhas if l[1] and str(l[1]).strip() == "Tipos ANBIMA")
    meses_col = {}   # (ano, mes) -> indice da coluna
    for indice, celula in enumerate(cabecalho):
        chave = parse_mes(celula)
        if chave is not None:
            meses_col[chave] = indice

    alvo = {sem_acento(r).strip().lower(): r for r in rotulos}
    linhas_tipo = {}
    for linha in linhas:
        rotulo = sem_acento(str(linha[1])).strip().lower() if linha[1] else ""
        if rotulo in alvo and alvo[rotulo] not in linhas_tipo:
            linhas_tipo[alvo[rotulo]] = linha

    faltando = set(rotulos) - set(linhas_tipo) - {"Ações Sustentabilidade / Governança"}
    if faltando:
        raise ValueError(f"Tipos nao encontrados em '{ws.title}': {sorted(faltando)}")

    dados = {}
    for chave in sorted(meses_col)[-n_meses:]:
        col = meses_col[chave]
        dados[chave] = {rot: (linha[col] if (linha := linhas_tipo.get(rot)) else None)
                        for rot in rotulos}
    return dados


def totais_anuais_pag8(ws):
    """Linhas de total anual (chave AAAA na coluna A) da Pag. 8."""
    totais = {}
    for linha in ws.iter_rows(values_only=True):
        valor = linha[0]
        if isinstance(valor, str) and valor.isdigit():
            valor = int(valor)
        if isinstance(valor, (int, float)) and not isinstance(valor, bool) and 1900 <= int(valor) <= 2999:
            totais[int(valor)] = linha
    return totais


# ---------------------------------------------------------------------------
# Upsert no historico
# ---------------------------------------------------------------------------

def upsert_mensal(linhas, dados, ordem_colunas):
    """Para abas com uma linha por mes (data na coluna A): substitui os valores
    dos meses presentes em `dados` e acrescenta ao final os meses novos, em
    ordem cronologica. `linhas` e a lista mutavel de listas da aba."""
    indice_por_mes = {}
    for i, linha in enumerate(linhas):
        chave = parse_mes(linha[0])
        if chave is not None:
            indice_por_mes[chave] = i

    for chave in sorted(dados):
        valores = [dados[chave].get(nome) for nome in ordem_colunas]
        if chave in indice_por_mes:
            linha = linhas[indice_por_mes[chave]]
            linha[1:1 + len(valores)] = valores
        else:
            linhas.append([datetime(chave[0], chave[1], 1)] + valores)


def upsert_cap_liq_classe(linhas, dados, totais_anuais, colunas_fonte, ordem_colunas):
    """Aba 'Cap. Liq. Por Classe': linhas anuais (rotulo AAAA) seguidas das
    linhas mensais daquele ano (rotulos 1..12). Substitui os meses da janela,
    acrescenta meses/anos novos e atualiza o total anual dos anos afetados."""
    def chave_ano(valor):
        if isinstance(valor, str) and valor.strip().isdigit():
            valor = int(valor.strip())
        if isinstance(valor, (int, float)) and not isinstance(valor, bool) and 1900 <= int(valor) <= 2999:
            return int(valor)
        return None

    def chave_mes(valor):
        if isinstance(valor, (int, float)) and not isinstance(valor, bool) and 1 <= int(valor) <= 12:
            return int(valor)
        if isinstance(valor, str) and valor.strip().isdigit() and 1 <= int(valor.strip()) <= 12:
            return int(valor.strip())
        return None

    # mapeia posicao de cada linha anual e de cada (ano, mes)
    posicao_ano, posicao_mes = {}, {}
    ano_corrente = None
    for i, linha in enumerate(linhas[1:], start=1):
        ano = chave_ano(linha[0])
        if ano is not None:
            ano_corrente = ano
            posicao_ano[ano] = i
            continue
        mes = chave_mes(linha[0])
        if mes is not None and ano_corrente is not None:
            posicao_mes[(ano_corrente, mes)] = i

    anos_afetados = sorted({ano for ano, _ in dados})
    for ano in anos_afetados:
        if ano not in posicao_ano:            # ano novo: abre bloco no final
            linhas.append([ano] + [None] * len(ordem_colunas))
            posicao_ano[ano] = len(linhas) - 1

    for (ano, mes) in sorted(dados):
        valores = [dados[(ano, mes)].get(nome) for nome in ordem_colunas]
        if (ano, mes) in posicao_mes:
            linha = linhas[posicao_mes[(ano, mes)]]
            linha[1:1 + len(valores)] = valores
        else:
            # insere apos a ultima linha do bloco do ano (total anual ou mes anterior)
            base = posicao_ano[ano]
            destino = base + 1
            while destino < len(linhas) and (m := chave_mes(linhas[destino][0])) is not None and m < mes:
                destino += 1
            linhas.insert(destino, [mes] + valores)
            posicao_ano = {a: (p + 1 if p >= destino else p) for a, p in posicao_ano.items()}
            posicao_mes = {c: (p + 1 if p >= destino else p) for c, p in posicao_mes.items()}
            posicao_mes[(ano, mes)] = destino

    for ano in anos_afetados:                 # atualiza o total anual (YTD para o ano corrente)
        linha_fonte = totais_anuais.get(ano)
        if linha_fonte is not None:
            valores = [linha_fonte[indice] for indice in colunas_fonte.values()]
            linha = linhas[posicao_ano[ano]]
            linha[1:1 + len(valores)] = valores


# ---------------------------------------------------------------------------
# Principal
# ---------------------------------------------------------------------------

TIPOS_ACOES = [
    "Ações", "Ações Indexados", "Ações Índice Ativo", "Ações Valor / Crescimento",
    "Ações Small Caps", "Ações Dividendos", "Ações Sustentabilidade / Governança",
    "Ações Setoriais", "Ações Livre", "Ações FMP-FGTS", "Fechados de Ações",
    "Fundo Mono Ação", "Ações Investimento no Exterior",
]

# indices (base 0) das colunas nas paginas da ANBIMA
COLS_PAG3 = {"Renda Fixa": 1, "Ações": 2, "Multimercados": 3, "Previdência": 5,
             "ETF": 6, "Total": 12}
COLS_PAG8 = {"Renda Fixa": 2, "Ações": 3, "Multimercados": 4, "Previdência": 6, "ETF": 7}


def carregar_linhas(ws):
    return [list(linha) for linha in ws.iter_rows(values_only=True)]


def main() -> None:
    parser = argparse.ArgumentParser(description="Consolida o relatorio mensal da ANBIMA com o historico proprio")
    parser.add_argument("--anbima", default=CAMINHO_ANBIMA, help="Planilha mensal da ANBIMA (Dados ambima.xlsx)")
    parser.add_argument("--historico", default=CAMINHO_HISTORICO, help="Base historica propria (Histórico.xlsx)")
    parser.add_argument("--saida", default=CAMINHO_SAIDA, help="Base consolidada de saida (Base XP.xlsx)")
    parser.add_argument("--meses", type=int, default=12, help="Janela de meses do relatorio ANBIMA a consolidar")
    args = parser.parse_args()

    wb_anbima = openpyxl.load_workbook(args.anbima, read_only=True, data_only=True)
    wb_hist = openpyxl.load_workbook(args.historico, data_only=True)

    pl_const = ultimos_meses_pag_classe(aba(wb_anbima, "Pág. 3 - PL Const. por Categ."), COLS_PAG3, args.meses)
    cap_classe = ultimos_meses_pag_classe(aba(wb_anbima, "Pág. 8 - Cap. Líq. por Classe"), COLS_PAG8, args.meses)
    totais_cap = totais_anuais_pag8(aba(wb_anbima, "Pág. 8 - Cap. Líq. por Classe"))
    pl_tipo = ultimos_meses_pag_tipo(aba(wb_anbima, "Pág. 5 - PL por Tipo"), TIPOS_ACOES, args.meses)
    cap_tipo = ultimos_meses_pag_tipo(aba(wb_anbima, "Pág. 9 - Cap. Líq. por Tipo"), TIPOS_ACOES, args.meses)
    wb_anbima.close()

    janela = sorted(pl_const)
    print(f"Janela consolidada: {janela[0][1]:02d}/{janela[0][0]} a {janela[-1][1]:02d}/{janela[-1][0]}")

    saida = openpyxl.Workbook()
    saida.remove(saida.active)

    especificacoes = [
        ("PL Const. Por Classe", pl_const,
         ["Renda Fixa", "Ações", "Multimercados", "Previdência", "ETF", "Total"], "mensal"),
        ("Cap. Líq. Por Classe", cap_classe,
         ["Renda Fixa", "Ações", "Multimercados", "Previdência", "ETF"], "cap_classe"),
        ("PL por Tipo", pl_tipo, TIPOS_ACOES, "mensal"),
        ("Cap. Líq. Tipo", cap_tipo, TIPOS_ACOES, "mensal"),
    ]
    for nome, dados, ordem, modo in especificacoes:
        ws_hist = aba(wb_hist, nome)
        linhas = carregar_linhas(ws_hist)
        if modo == "mensal":
            upsert_mensal(linhas, dados, ordem)
        else:
            upsert_cap_liq_classe(linhas, dados, totais_cap, COLS_PAG8, ordem)

        ws_saida = saida.create_sheet(title=ws_hist.title)
        for linha in linhas:
            ws_saida.append(linha)
        for celula in ws_saida["A"]:
            if isinstance(celula.value, datetime):
                celula.number_format = "mmm-yy"
        print(f"  {nome}: {len(linhas) - 1} linhas de dados")

    saida.save(args.saida)
    print(f"Base consolidada salva em: {args.saida}")


if __name__ == "__main__":
    main()
