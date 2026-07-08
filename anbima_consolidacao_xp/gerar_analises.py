"""Gera as tabelas de analise a partir da Base XP.xlsx consolidada.

Saida: um arquivo Excel com 5 abas:
  1. 'PL por Classe (12M)'          - PL de Renda Fixa, Acoes e Multimercados nos ultimos 12 meses
  2. 'Cap. Liq. por Classe (12M)'   - captacao liquida mensal por classe nos ultimos 12 meses
  3. 'AUM Allocation (Annual)'      - dezembro de cada ano + ultimo mes disponivel (ano corrente):
        Date | Fixed Income as % of Total Net AUM | Equities as % of Total Net AUM (RHS) |
        Hedge Funds as % of Total Net AUM | Others as % of Total Net AUM |
        Total Net AUM | Total equities AUM | Total AUM of funds industry
        (percentuais sobre o PL total da industria; Total Net AUM = RF + Acoes + MM)
  4. 'Aloc. Acoes por Tipo (Anual)' - % de cada tipo de fundo de acoes no PL total de Acoes,
        dezembro de cada ano desde 2016 + ultimo mes disponivel
  5. 'Cap. Liq. Acoes Tipo (12M)'   - captacao liquida mensal por tipo de fundo de acoes (LTM)

Uso:
    python gerar_analises.py
    python gerar_analises.py --base "caminho\\Base XP.xlsx" --saida "caminho\\Análises.xlsx"
"""
import argparse
from datetime import datetime

import openpyxl

CAMINHO_BASE = r"\\xpdocs\Research\Equities\Estrategia\Reports\Fluxo investidores na Bolsa\Banco de dados\Anbima Automation\input\Base XP.xlsx"
CAMINHO_SAIDA = r"\\xpdocs\Research\Equities\Estrategia\Reports\Fluxo investidores na Bolsa\Banco de dados\Anbima Automation\input\Análises.xlsx"

FORMATO_VALOR = "#,##0"
FORMATO_PCT = "0.0%"
FORMATO_DATA = "mmm-yy"


def num(valor):
    return valor if isinstance(valor, (int, float)) and not isinstance(valor, bool) else 0


def linhas_mensais(ws):
    """Linhas cuja coluna A e uma data, na ordem da planilha."""
    return [list(r) for r in ws.iter_rows(values_only=True) if isinstance(r[0], datetime)]


def cap_liq_classe_mensal(ws):
    """Converte a aba 'Cap. Líq. Por Classe' (linha de ano seguida dos meses
    1..12) em linhas mensais [datetime, RF, Acoes, MM, Prev, ETF]."""
    mensal = []
    ano_corrente = None
    for r in ws.iter_rows(values_only=True):
        chave = r[0]
        if isinstance(chave, str) and chave.strip().isdigit():
            chave = int(chave.strip())
        if isinstance(chave, (int, float)) and not isinstance(chave, bool):
            chave = int(chave)
            if 1900 <= chave <= 2999:
                ano_corrente = chave
            elif 1 <= chave <= 12 and ano_corrente is not None:
                mensal.append([datetime(ano_corrente, chave, 1)] + list(r[1:6]))
    return mensal


def escrever_aba(wb, titulo, cabecalho, linhas, formatos):
    """Cria uma aba com cabecalho e aplica formato por coluna ('data', 'valor',
    'pct' ou None)."""
    ws = wb.create_sheet(title=titulo)
    ws.append(cabecalho)
    for c in ws[1]:
        c.font = openpyxl.styles.Font(bold=True)
    for linha in linhas:
        ws.append(linha)
    mapa = {"data": FORMATO_DATA, "valor": FORMATO_VALOR, "pct": FORMATO_PCT}
    for j, tipo in enumerate(formatos, start=1):
        if tipo in mapa:
            for i in range(2, ws.max_row + 1):
                ws.cell(row=i, column=j).number_format = mapa[tipo]
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 12
    return ws


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera as tabelas de analise da Base XP")
    parser.add_argument("--base", default=CAMINHO_BASE, help="Base consolidada (Base XP.xlsx)")
    parser.add_argument("--saida", default=CAMINHO_SAIDA, help="Arquivo de saida com as analises")
    parser.add_argument("--meses", type=int, default=12, help="Janela LTM das tabelas mensais")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.base, data_only=True)
    pl_classe = linhas_mensais(wb["PL Const. Por Classe"])       # Periodo, RF, Acoes, MM, Prev, ETF, Total
    cap_classe = cap_liq_classe_mensal(wb["Cap. Líq. Por Classe"])
    pl_tipo = linhas_mensais(wb["PL por Tipo"])                  # Data, Acoes, 12 subtipos
    cap_tipo = linhas_mensais(wb["Cap. Líq. Tipo"])
    cab_tipos = [r for r in wb["PL por Tipo"].iter_rows(values_only=True)][0]
    subtipos = list(cab_tipos[2:])                               # nomes dos 12 tipos de acoes
    wb.close()

    saida = openpyxl.Workbook()
    saida.remove(saida.active)

    # 1. PL por 3 classes - ultimos 12 meses
    tabela1 = [[r[0], r[1], r[2], r[3]] for r in pl_classe[-args.meses:]]
    escrever_aba(saida, "PL por Classe (12M)",
                 ["Período", "Renda Fixa", "Ações", "Multimercados"],
                 tabela1, ["data", "valor", "valor", "valor"])

    # 2. Captacao liquida por classe - ultimos 12 meses
    tabela2 = cap_liq = cap_classe[-args.meses:]
    escrever_aba(saida, "Cap. Líq. por Classe (12M)",
                 ["Período", "Renda Fixa", "Ações", "Multimercados", "Previdência", "ETF"],
                 tabela2, ["data"] + ["valor"] * 5)

    # 3. Alocacao anual (dezembro de cada ano + ultimo mes como ano corrente)
    anuais = [r for r in pl_classe if r[0].month == 12]
    ultimo = pl_classe[-1]
    if ultimo[0].month != 12:
        anuais.append(ultimo)
    tabela3 = []
    for r in anuais:
        data, rf, eq, hf, total = r[0], num(r[1]), num(r[2]), num(r[3]), num(r[6])
        tabela3.append([
            data,
            rf / total if total else None,
            eq / total if total else None,
            hf / total if total else None,
            (total - rf - eq - hf) / total if total else None,
            rf + eq + hf,
            eq,
            total,
        ])
    escrever_aba(saida, "AUM Allocation (Annual)",
                 ["Date", "Fixed Income as % of Total Net AUM", "Equities as % of Total Net AUM (RHS)",
                  "Hedge Funds as % of Total Net AUM", "Others as % of Total Net AUM",
                  "Total Net AUM", "Total equities AUM", "Total AUM of funds industry"],
                 tabela3, ["data", "pct", "pct", "pct", "pct", "valor", "valor", "valor"])

    # 4. Alocacao relativa por tipo de fundo de acoes (anual, desde 2016)
    anuais_tipo = [r for r in pl_tipo if r[0].month == 12 and r[0].year >= 2016]
    ultimo = pl_tipo[-1]
    if ultimo[0].month != 12:
        anuais_tipo.append(ultimo)
    tabela4 = []
    for r in anuais_tipo:
        total_acoes = num(r[1])
        pesos = [(num(v) / total_acoes if total_acoes and v is not None else None) for v in r[2:]]
        tabela4.append([r[0]] + pesos + [total_acoes])
    escrever_aba(saida, "Aloc. Ações por Tipo (Anual)",
                 ["Date"] + subtipos + ["Total Ações (R$ mi)"],
                 tabela4, ["data"] + ["pct"] * len(subtipos) + ["valor"])

    # 5. Captacao liquida por tipo de fundo de acoes - mensal LTM
    tabela5 = cap_tipo[-args.meses:]
    escrever_aba(saida, "Cap. Líq. Ações Tipo (12M)",
                 ["Data", "Ações (total)"] + subtipos,
                 tabela5, ["data"] + ["valor"] * (len(subtipos) + 1))

    saida.save(args.saida)
    print(f"Análises salvas em: {args.saida}")
    print(f"  1. PL por Classe (12M): {len(tabela1)} meses ({tabela1[0][0]:%m/%Y} a {tabela1[-1][0]:%m/%Y})")
    print(f"  2. Cap. Líq. por Classe (12M): {len(tabela2)} meses")
    print(f"  3. AUM Allocation (Annual): {len(tabela3)} linhas ({tabela3[0][0]:%Y} a {tabela3[-1][0]:%m/%Y})")
    print(f"  4. Aloc. Ações por Tipo (Anual): {len(tabela4)} linhas")
    print(f"  5. Cap. Líq. Ações por Tipo (12M): {len(tabela5)} meses")


if __name__ == "__main__":
    main()
