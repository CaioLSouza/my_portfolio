"""Anexa uma nova linha (data + valor) no fim das planilhas historicas de
pessoa fisica na Bolsa, preservando o conteudo e a formatacao existentes.

Usa openpyxl (e nao pandas.to_excel) de proposito: to_excel reescreveria o
arquivo inteiro e perderia formatacao, formulas e graficos. Aqui a gente so
acrescenta uma linha ao final, copiando o formato numerico da celula de cima.

A operacao e idempotente: se a data ja existir na planilha, a linha nao e
duplicada (por padrao apenas avisa; com sobrescrever=True atualiza o valor).
"""
from __future__ import annotations

import shutil
from datetime import date, datetime
from pathlib import Path


def _para_data(valor) -> date | None:
    """Normaliza o conteudo de uma celula de data para `date`, aceitando tanto
    datetime (como o Excel guarda datas) quanto texto dd/mm/aaaa."""
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    return None


def _achar_coluna(planilha, nome_coluna: str, linha_cabecalho: int = 1) -> int:
    """Descobre o indice (1-based) da coluna cujo cabecalho casa com `nome_coluna`
    (comparacao sem diferenciar maiuscula/minuscula e espacos nas bordas)."""
    alvo = nome_coluna.strip().lower()
    for celula in planilha[linha_cabecalho]:
        if celula.value is not None and str(celula.value).strip().lower() == alvo:
            return celula.column
    disponiveis = [c.value for c in planilha[linha_cabecalho] if c.value is not None]
    raise ValueError(
        f"Coluna '{nome_coluna}' nao encontrada na aba '{planilha.title}'. "
        f"Cabecalhos encontrados: {disponiveis}"
    )


def _ultima_linha_com_dados(planilha, coluna_data: int, linha_cabecalho: int) -> int:
    """Ultima linha que tem data preenchida (max_row do openpyxl pode contar
    linhas vazias com formatacao residual)."""
    ultima = linha_cabecalho
    for linha in range(linha_cabecalho + 1, planilha.max_row + 1):
        if planilha.cell(row=linha, column=coluna_data).value not in (None, ""):
            ultima = linha
    return ultima


def anexar_valor(
    caminho: str | Path,
    valor,
    data_referencia: datetime | date,
    coluna_valor: str,
    coluna_data: str = "data",
    aba: str | None = None,
    linha_cabecalho: int = 1,
    dry_run: bool = False,
    fazer_backup: bool = True,
    sobrescrever: bool = False,
) -> str:
    """Anexa (data_referencia, valor) na planilha. Devolve uma mensagem
    descrevendo o que foi (ou seria) feito."""
    from openpyxl import load_workbook

    caminho = Path(caminho)
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {caminho}")

    data_ref = data_referencia.date() if isinstance(data_referencia, datetime) else data_referencia

    livro = load_workbook(caminho)
    planilha = livro[aba] if aba else livro.worksheets[0]

    col_data = _achar_coluna(planilha, coluna_data, linha_cabecalho)
    col_valor = _achar_coluna(planilha, coluna_valor, linha_cabecalho)
    ultima = _ultima_linha_com_dados(planilha, col_data, linha_cabecalho)

    # A data ja existe? Evita duplicar o mesmo mes/dia.
    for linha in range(linha_cabecalho + 1, ultima + 1):
        if _para_data(planilha.cell(row=linha, column=col_data).value) == data_ref:
            if not sobrescrever:
                return (
                    f"[ignorado] {caminho.name}: data {data_ref:%d/%m/%Y} ja existe "
                    f"na linha {linha}; use sobrescrever=True para atualizar."
                )
            if not dry_run:
                planilha.cell(row=linha, column=col_valor).value = valor
                livro.save(caminho)
            return (
                f"[atualizado] {caminho.name}: linha {linha} "
                f"({data_ref:%d/%m/%Y}) -> {coluna_valor}={valor}"
            )

    destino = ultima + 1
    if dry_run:
        return (
            f"[dry-run] {caminho.name}: adicionaria na linha {destino} "
            f"{coluna_data}={data_ref:%d/%m/%Y}, {coluna_valor}={valor}"
        )

    if fazer_backup:
        shutil.copyfile(caminho, caminho.with_suffix(".bak" + caminho.suffix))

    celula_data = planilha.cell(row=destino, column=col_data, value=data_ref)
    celula_valor = planilha.cell(row=destino, column=col_valor, value=valor)
    # Herda o formato numerico da celula de cima para manter a aparencia.
    if destino - 1 > linha_cabecalho:
        celula_data.number_format = planilha.cell(row=destino - 1, column=col_data).number_format
        celula_valor.number_format = planilha.cell(row=destino - 1, column=col_valor).number_format

    livro.save(caminho)
    return (
        f"[ok] {caminho.name}: linha {destino} "
        f"{coluna_data}={data_ref:%d/%m/%Y}, {coluna_valor}={valor}"
    )
