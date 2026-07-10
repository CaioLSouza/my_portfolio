"""Teste rapido da logica pura (parsing e escrita em Excel), sem Selenium/B3."""
import io
import tempfile
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

import planilhas
import scraper_b3_pf as s

print("### 1) parsing de numeros brasileiros")
casos = {"1.234.567": 1234567.0, "12.345.678,90": 12345678.90, "R$ 1.500,25": 1500.25,
         "45,7%": 45.7, "  ": None, "-": None, "5000000": 5000000.0}
for entrada, esperado in casos.items():
    got = s._para_numero(entrada)
    assert got == esperado, f"{entrada!r} -> {got} (esperava {esperado})"
print("  ok")

print("### 2) extrair_dados COM linha de total")
html_total = """
<table>
 <tr><th>Faixa etária</th><th>Quantidade de contas</th><th>% Qtde</th><th>Valor (R$)</th><th>% Valor</th></tr>
 <tr><td>Até 15 anos</td><td>100.000</td><td>2%</td><td>1.000.000.000,00</td><td>1%</td></tr>
 <tr><td>16 a 25 anos</td><td>900.000</td><td>18%</td><td>9.000.000.000,00</td><td>9%</td></tr>
 <tr><td>Total</td><td>5.000.000</td><td>100%</td><td>500.000.000.000,00</td><td>100%</td></tr>
</table>
"""
tabelas = pd.read_html(io.StringIO(html_total))
num, pos, origem = s.extrair_dados(tabelas)
assert num == 5_000_000 and pos == 500_000_000_000.0 and origem == "linha de total", (num, pos, origem)
print(f"  ok -> contas={num:,.0f} posicao={pos:,.0f} ({origem})")

print("### 3) extrair_dados SEM linha de total (soma das faixas)")
html_sem = """
<table>
 <tr><th>Faixa etária</th><th>Nº de investidores</th><th>Posição em custódia (R$)</th></tr>
 <tr><td>16 a 25 anos</td><td>2.000.000</td><td>50.000.000.000,00</td></tr>
 <tr><td>26 a 35 anos</td><td>3.000.000</td><td>150.000.000.000,00</td></tr>
</table>
"""
num2, pos2, origem2 = s.extrair_dados(pd.read_html(io.StringIO(html_sem)))
assert num2 == 5_000_000 and pos2 == 200_000_000_000.0 and origem2 == "soma das faixas", (num2, pos2, origem2)
print(f"  ok -> contas={num2:,.0f} posicao={pos2:,.0f} ({origem2})")

print("### 4) escrita no Excel (append + dedupe + dry-run)")
caminho = Path(tempfile.gettempdir()) / "_pf_teste.xlsx"
wb = Workbook(); ws = wb.active
ws.append(["data", "Número"])
ws.append([datetime(2024, 1, 31), 4_000_000])
ws.cell(row=2, column=1).number_format = "DD/MM/YYYY"
wb.save(caminho)

msg = planilhas.anexar_valor(caminho, 5_000_000, datetime(2024, 2, 29), "Número")
print("  ", msg); assert msg.startswith("[ok]")
wb2 = load_workbook(caminho)
assert wb2.active.cell(row=3, column=2).value == 5_000_000
assert wb2.active.cell(row=3, column=1).value.date().isoformat() == "2024-02-29"
assert wb2.active.cell(row=3, column=1).number_format == "DD/MM/YYYY", "formato de data nao herdado"

msg_dup = planilhas.anexar_valor(caminho, 9, datetime(2024, 2, 29), "Número")
print("  ", msg_dup); assert msg_dup.startswith("[ignorado]")

msg_dry = planilhas.anexar_valor(caminho, 6_000_000, datetime(2024, 3, 31), "Número", dry_run=True)
print("  ", msg_dry); assert msg_dry.startswith("[dry-run]")
assert load_workbook(caminho).active.max_row == 3, "dry-run nao deveria ter gravado"

msg_over = planilhas.anexar_valor(caminho, 5_555_555, datetime(2024, 2, 29), "Número", sobrescrever=True)
print("  ", msg_over); assert msg_over.startswith("[atualizado]")
assert load_workbook(caminho).active.cell(row=3, column=2).value == 5_555_555
print("  ok")

print("\nTODOS OS TESTES PASSARAM")
