# -*- coding: utf-8 -*-
"""
Migra o historico da PA_Principal.xlsx para a BASE MESTRE hibrida.

Funciona de DUAS formas:

  1. Interactive Window / Jupyter (VS Code):
     ajuste os caminhos na celula [1] e rode as celulas em ordem
     (Shift+Enter), ou "Run All". Tambem da para chamar direto:
         info = migrar("PA_Principal.xlsx", "PA_Base_Historica.xlsx")

  2. Linha de comando:
         python3 migrar_base.py PA_Principal.xlsx -o PA_Base_Historica.xlsx

Saida (PA_Base_Historica.xlsx):
    - "Raw Data"  : copia fiel da aba larga atual (canonica)
    - "Respostas" : visao LONG (1 linha por resposta x opcao; ';' explodido)
    - "Perguntas" : catalogo com primeira/ultima edicao e flag ATIVA
    - "LEIA-ME"   : instrucoes
"""

# %% [1] Parametros (edite aqui quando rodar no Interactive Window) -----------
ARQUIVO_ORIGEM = "PA_Principal.xlsx"          # planilha atual, com a aba "Raw Data"
ARQUIVO_SAIDA = "PA_Base_Historica.xlsx"      # base mestre a ser gerada

# %% [2] Imports e constantes -------------------------------------------------
import sys
import openpyxl
from openpyxl.utils import get_column_letter as gcl
from openpyxl.styles import Font, PatternFill

META_COLS = 5          # A..E: periodo, inicio, conclusao, email, nome
PERIOD_COL = 1

NAVY = "1F2F44"
WHITE = "FFFFFF"


# %% [3] Funcao de migracao ----------------------------------------------------
def migrar(origem, destino):
    src = openpyxl.load_workbook(origem, data_only=True)
    raw = src["Raw Data"]

    # ---- limites reais ----
    last_row = 0
    for r in range(2, raw.max_row + 1):
        if raw.cell(r, PERIOD_COL).value is not None:
            last_row = r
    last_col = 0
    for c in range(1, raw.max_column + 1):
        if raw.cell(1, c).value is not None:
            last_col = c

    out = openpyxl.Workbook()
    out.remove(out.active)

    # ================= 1) Raw Data (copia fiel, so valores) =================
    w = out.create_sheet("Raw Data")
    for r in range(1, last_row + 1):
        for c in range(1, last_col + 1):
            v = raw.cell(r, c).value
            if v is not None:
                w.cell(r, c, v)
    for c in range(1, last_col + 1):
        w.cell(1, c).font = Font(bold=True, size=9, color=WHITE)
        w.cell(1, c).fill = PatternFill("solid", fgColor=NAVY)
    w.freeze_panes = "A2"

    # ================= 2) Respostas (long explodida) ========================
    lg = out.create_sheet("Respostas")
    hdr = ["periodo", "id_resposta", "pergunta", "opcao"]
    for j, h in enumerate(hdr, 1):
        cell = lg.cell(1, j, h)
        cell.font = Font(bold=True, size=9, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
    lr = 2
    stats = {}   # pergunta -> {first, last, n, col}
    for r in range(2, last_row + 1):
        periodo = raw.cell(r, PERIOD_COL).value
        if periodo is None:
            continue
        for c in range(META_COLS + 1, last_col + 1):
            v = raw.cell(r, c).value
            if v is None or v == "":
                continue
            # texto EXATO do cabecalho (com nbsp etc.): e por ele que o
            # Power Query seleciona colunas e os MATCH da Base casam
            pergunta = str(raw.cell(1, c).value or "")
            if not pergunta.strip():
                continue
            s = str(v)
            partes = [p.strip() for p in s.split(";")] if ";" in s else [s.strip()]
            for p in partes:
                if not p:
                    continue
                lg.cell(lr, 1, periodo)
                lg.cell(lr, 2, r)          # id = linha original no Raw Data
                lg.cell(lr, 3, pergunta)
                lg.cell(lr, 4, p)
                lr += 1
            st = stats.setdefault(pergunta, {"first": periodo, "last": periodo, "n": 0, "col": c})
            st["first"] = min(st["first"], periodo)
            st["last"] = max(st["last"], periodo)
            st["n"] += 1
    lg.freeze_panes = "A2"
    lg.column_dimensions["C"].width = 70
    lg.column_dimensions["D"].width = 45

    # ================= 3) Perguntas (catalogo + flag Ativa) ==================
    ultimo_periodo = max(st["last"] for st in stats.values())
    cat = out.create_sheet("Perguntas")
    hdr = ["col_raw", "pergunta", "primeira_edicao", "ultima_edicao",
           "n_respostas_total", "ativa"]
    for j, h in enumerate(hdr, 1):
        cell = cat.cell(1, j, h)
        cell.font = Font(bold=True, size=9, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
    r = 2
    for pergunta, st in sorted(stats.items(), key=lambda kv: kv[1]["col"]):
        cat.cell(r, 1, gcl(st["col"]))
        cat.cell(r, 2, pergunta)
        cat.cell(r, 3, st["first"])
        cat.cell(r, 4, st["last"])
        cat.cell(r, 5, st["n"])
        ativa = 1 if st["last"] == ultimo_periodo else 0
        c6 = cat.cell(r, 6, ativa)
        if ativa:
            c6.fill = PatternFill("solid", fgColor="C6EFCE")
        r += 1
    cat.column_dimensions["B"].width = 80
    cat.freeze_panes = "A2"

    # ================= 4) LEIA-ME ============================================
    rd = out.create_sheet("LEIA-ME")
    notas = [
        "BASE MESTRE - Pesquisa XP com assessores (historico completo)",
        "",
        "Este arquivo e o repositorio historico. Regras:",
        "  1. APPEND-ONLY: so se adiciona dados (via macro de importacao do Forms).",
        "     Nunca editar/apagar linhas antigas.",
        "  2. 'Raw Data' e a visao canonica (larga). 'Respostas' e a visao long",
        "     derivada (1 linha por resposta x opcao) - usada pelo Power Query.",
        "  3. 'Perguntas' e o catalogo. A coluna ATIVA (0/1) controla quais",
        "     perguntas a planilha de producao carrega. Aposentou uma pergunta?",
        "     Troque ativa para 0 - o historico permanece intacto aqui.",
        "",
        "A planilha de producao (a que gera o relatorio) NAO guarda historico:",
        "ela puxa desta base via Power Query (ultimos N meses x perguntas ativas).",
    ]
    for i, n in enumerate(notas, 1):
        rd.cell(i, 1, n).font = Font(bold=(i == 1), size=11)
    rd.column_dimensions["A"].width = 90

    out.save(destino)
    return {
        "linhas_raw": last_row - 1,
        "colunas_raw": last_col,
        "linhas_long": lr - 2,
        "perguntas": len(stats),
        "ativas": sum(1 for st in stats.values() if st["last"] == ultimo_periodo),
        "ultimo_periodo": ultimo_periodo,
    }


def _interativo():
    """True quando rodando em Jupyter/Interactive Window (VS Code)."""
    return "ipykernel" in sys.modules or hasattr(sys, "ps1")


# %% [4] Executar ---------------------------------------------------------------
if __name__ == "__main__":
    if _interativo():
        # Interactive Window: usa os parametros da celula [1]
        info = migrar(ARQUIVO_ORIGEM, ARQUIVO_SAIDA)
        print("Base mestre gerada:", ARQUIVO_SAIDA)
        for k, v in info.items():
            print(f"  {k}: {v}")
    else:
        # Linha de comando: argumentos via argparse
        import argparse
        ap = argparse.ArgumentParser()
        ap.add_argument("origem", nargs="?", default=ARQUIVO_ORIGEM)
        ap.add_argument("-o", "--out", default=ARQUIVO_SAIDA)
        args = ap.parse_args()
        info = migrar(args.origem, args.out)
        print("Base mestre gerada:", args.out)
        for k, v in info.items():
            print(f"  {k}: {v}")
